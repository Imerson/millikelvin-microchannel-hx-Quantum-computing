#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

BLUE=Font(name="Arial",color="0000FF"); BLK=Font(name="Arial"); BOLD=Font(name="Arial",bold=True)
HDR=Font(name="Arial",bold=True,color="FFFFFF"); HFILL=PatternFill("solid",fgColor="1F3A5F")
NOTE=Font(name="Arial",italic=True,size=9,color="555555")
thin=Side(style="thin",color="BFBFBF"); BORD=Border(thin,thin,thin,thin)
def hdr(ws,row,cells):
    for c,v in cells.items():
        ws[f"{c}{row}"]=v; ws[f"{c}{row}"].font=HDR; ws[f"{c}{row}"].fill=HFILL
        ws[f"{c}{row}"].alignment=Alignment(horizontal="center",wrap_text=True); ws[f"{c}{row}"].border=BORD

wb=Workbook()

# ---------------- INPUTS ----------------
ws=wb.active; ws.title="Inputs"
ws["A1"]="Sinter vs Microchannel — inputs and derived quantities"; ws["A1"].font=Font(name="Arial",bold=True,size=13)
rows=[("Block length L (m)",0.100,"geometry, normalised block"),
      ("Block width W (m)",0.040,""),("Block height H (m)",0.040,""),
      ("Face area A_face (m^2)","=B4*B5","W*H"),
      ("He-4 viscosity mu (Pa.s)",1.0e-6,"He-II normal-fluid value (Pobell 2007)"),
      ("He-4 density rho (kg/m^3)",145.0,"Donnelly & Barenghi 1998"),
      ("Interface temperature T_ref (K)",0.025,"CFD conjugate interface temperature"),
      ("Kapitza coeff C_K Cu (K^4 m^2/W)",0.020,"Pollack 1969"),
      ("Kapitza coeff C_K Ag (K^4 m^2/W)",0.005,"Dransfeld & Salzmann 1967"),
      ("Microchannel velocity U_ch (m/s)",1.0e-3,"per-channel inlet velocity (CFD)"),
      ("Channel open-area fraction","=1/1.5^2","pitch = 1.5 Dh"),
      ("Superficial velocity Us (m/s)","=B13*B12","open_fraction * U_ch (common basis)"),
      ("Volumetric flow Vdot (m^3/s)","=B14*B6","Us * A_face (same for all geometries)"),
      ("Sinter particle dp (m)",0.07e-6,"Nakagawa 2023 (0.07 um silver powder)"),
      ("Sinter porosity eps",0.50,"standard pressed Ag sinter (Pobell 2007); paper does not state"),
      ("Permeability k (m^2)","=B17^3*B16^2/(180*(1-B17)^2)","Kozeny-Carman"),
      ("Specific area a_v (m^2/m^3)","=6*(1-B17)/B16","packed spheres"),
      ("Idealised sinter area (m^2)","=B19*B6*B3","a_v * V_block (full microscopic area)"),
      ("Sinter Darcy dP (Pa)","=B7*B14*B3/B18","mu*Us*L/k"),
      ("Pore Reynolds Re_pore","=B8*B14*B16/B7","<<1 => Darcy valid, Ergun term negligible")]
r=3
for lab,val,note in rows:
    ws[f"A{r}"]=lab; ws[f"A{r}"].font=BLK
    ws[f"B{r}"]=val
    ws[f"B{r}"].font=(BLUE if not (isinstance(val,str) and val.startswith("=")) else BLK)
    ws[f"B{r}"].number_format="0.000E+00" if isinstance(val,(int,float)) and (abs(val)<0.01 or abs(val)>=1000) else "0.0000"
    ws[f"C{r}"]=note; ws[f"C{r}"].font=NOTE
    r+=1
ws.column_dimensions["A"].width=34; ws.column_dimensions["B"].width=16; ws.column_dimensions["C"].width=48
ws["A24"]="Blue = input ; black = formula."; ws["A24"].font=NOTE

# ---------------- MICROCHANNEL ----------------
ws=wb.create_sheet("Microchannel")
ws["A1"]="Microchannel cases (Delta P from converged CFD)"; ws["A1"].font=Font(name="Arial",bold=True,size=12)
hdr(ws,3,{"A":"Case","B":"Material","C":"Dh (mm)","D":"N_ch","E":"A_wet (m^2)","F":"dP CFD (Pa)",
          "G":"C_K","H":"R_total (K/W)","I":"G (W/K)","J":"W_pump (W)","K":"FOM (1/K)"})
mc=[("Cu_0p5","Cu",0.5,2809,0.5618,1.137e-2),("Cu_1p0","Cu",1.0,676,0.2704,2.923e-3),
    ("Cu_2p0","Cu",2.0,169,0.1352,8.078e-4),("Ag_0p5","Ag",0.5,2809,0.5618,1.137e-2),
    ("Ag_1p0","Ag",1.0,676,0.2704,2.923e-3),("Ag_2p0","Ag",2.0,169,0.1352,8.078e-4)]
r=4
for case,mat,Dh,Nch,Aw,dP in mc:
    ws[f"A{r}"]=case; ws[f"B{r}"]=mat; ws[f"C{r}"]=Dh; ws[f"D{r}"]=Nch; ws[f"E{r}"]=Aw; ws[f"F{r}"]=dP
    for c in "EF": ws[f"{c}{r}"].font=BLUE
    for c in "CD": ws[f"{c}{r}"].font=BLUE
    ws[f"G{r}"]=f'=IF(B{r}="Cu",Inputs!$B$10,Inputs!$B$11)'
    ws[f"H{r}"]=f"=G{r}/(Inputs!$B$9^3*E{r})"
    ws[f"I{r}"]=f"=1/H{r}"
    ws[f"J{r}"]=f"=F{r}*Inputs!$B$15"
    ws[f"K{r}"]=f"=I{r}/J{r}"
    for c in "HIJK": ws[f"{c}{r}"].number_format="0.000E+00"; ws[f"{c}{r}"].font=BLK
    ws[f"E{r}"].number_format="0.0000"; ws[f"F{r}"].number_format="0.000E+00"
    for c in "ABCDEFGHIJK": ws[f"{c}{r}"].border=BORD
    r+=1
for col,w in zip("ABCDEFGHIJK",[10,9,8,8,12,12,9,13,12,12,12]): ws.column_dimensions[col].width=w

# ---------------- SINTER ----------------
ws=wb.create_sheet("Sinter")
ws["A1"]="Sinter cases (Darcy + Kozeny-Carman; area = packed-sphere)"; ws["A1"].font=Font(name="Arial",bold=True,size=12)
hdr(ws,3,{"A":"Case","B":"Material","C":"Area basis","D":"A_wet (m^2)","E":"dP (Pa)",
          "F":"C_K","G":"R_total (K/W)","H":"G (W/K)","I":"W_pump (W)","J":"FOM (1/K)"})
sin=[("Sinter Ag (idealised)","Ag","full microscopic","=Inputs!$B$20"),
     ("Sinter Cu (idealised)","Cu","full microscopic","=Inputs!$B$20"),
     ("Sinter Ag (He-4 effective)","Ag","geometric x0.001","=Inputs!$B$20*0.001"),
     ("Sinter Cu (He-4 effective)","Cu","geometric x0.001","=Inputs!$B$20*0.001")]
r=4
for case,mat,basis,Aexpr in sin:
    ws[f"A{r}"]=case; ws[f"B{r}"]=mat; ws[f"C{r}"]=basis; ws[f"D{r}"]=Aexpr
    ws[f"E{r}"]="=Inputs!$B$21"
    ws[f"F{r}"]=f'=IF(B{r}="Cu",Inputs!$B$10,Inputs!$B$11)'
    ws[f"G{r}"]=f"=F{r}/(Inputs!$B$9^3*D{r})"
    ws[f"H{r}"]=f"=1/G{r}"
    ws[f"I{r}"]=f"=E{r}*Inputs!$B$15"
    ws[f"J{r}"]=f"=H{r}/I{r}"
    for c in "DEGHIJ": ws[f"{c}{r}"].number_format="0.000E+00"; ws[f"{c}{r}"].font=BLK
    for c in "ABCDEFGHIJ": ws[f"{c}{r}"].border=BORD
    r+=1
ws[f"A{r+1}"]="He-4 effective row applies Nakagawa 2023: only the geometrical (not microscopic) area is thermally effective for superfluid He-4; the 0.001 factor is illustrative — the conclusion holds for any factor <1."
ws[f"A{r+1}"].font=NOTE
for col,w in zip("ABCDEFGHIJ",[26,9,16,12,12,9,13,12,12,12]): ws.column_dimensions[col].width=w

# ---------------- COMPARISON ----------------
ws=wb.create_sheet("Comparison")
ws["A1"]="Comparison — figure of merit = thermal conductance / pumping power"; ws["A1"].font=Font(name="Arial",bold=True,size=12)
hdr(ws,3,{"A":"Geometry / case","B":"R_total (K/W)","C":"G (W/K)","D":"dP (Pa)","E":"W_pump (W)","F":"FOM (1/K)"})
ref=[("Microchannel Ag, 0.5 mm","Microchannel!H7","Microchannel!I7","Microchannel!F7","Microchannel!J7","Microchannel!K7"),
     ("Microchannel Cu, 0.5 mm","Microchannel!H4","Microchannel!I4","Microchannel!F4","Microchannel!J4","Microchannel!K4"),
     ("Microchannel Ag, 2.0 mm","Microchannel!H9","Microchannel!I9","Microchannel!F9","Microchannel!J9","Microchannel!K9"),
     ("Sinter Ag (idealised)","Sinter!G4","Sinter!H4","Sinter!E4","Sinter!I4","Sinter!J4"),
     ("Sinter Cu (idealised)","Sinter!G5","Sinter!H5","Sinter!E5","Sinter!I5","Sinter!J5"),
     ("Sinter Ag (He-4 effective)","Sinter!G6","Sinter!H6","Sinter!E6","Sinter!I6","Sinter!J6")]
r=4
for lab,Rk,G,dP,Wp,F in ref:
    ws[f"A{r}"]=lab
    ws[f"B{r}"]=f"={Rk}"; ws[f"C{r}"]=f"={G}"; ws[f"D{r}"]=f"={dP}"; ws[f"E{r}"]=f"={Wp}"; ws[f"F{r}"]=f"={F}"
    for c in "BCDEF": ws[f"{c}{r}"].number_format="0.000E+00"; ws[f"{c}{r}"].font=Font(name="Arial",color="008000")
    for c in "ABCDEF": ws[f"{c}{r}"].border=BORD
    r+=1
ws[f"A{r+1}"]="Headline FOM ratio (microchannel Ag 0.5 mm / idealised Ag sinter):"
ws[f"F{r+1}"]="=Microchannel!K7/Sinter!J4"; ws[f"F{r+1}"].font=BOLD; ws[f"F{r+1}"].number_format="0"
ws[f"A{r+3}"]=("Notes: (1) compared on a normalised 40x40x100 mm block, same He-4 coolant, same superficial velocity. "
              "(2) The FOM ratio is independent of velocity and temperature (both scale identically). "
              "(3) Sinter wins raw conductance; microchannel wins conductance-per-pumping by ~4 orders because open channels "
              "have negligible flow resistance while fine sinter pores impose ~MPa Darcy pressure drop. "
              "(4) For He-4 the sinter effective area is far below its microscopic area (Nakagawa 2023), widening the gap.")
ws[f"A{r+3}"].font=NOTE; ws.row_dimensions[r+3].height=70
ws.merge_cells(f"A{r+3}:F{r+3}"); ws[f"A{r+3}"].alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABCDEF",[30,14,12,12,12,14]): ws.column_dimensions[col].width=w

# ---------------- CHARTS ----------------
ws=wb.create_sheet("Charts")
for i,(img,cell) in enumerate([("fig_FOM.png","A1"),("fig_dP.png","A24"),("fig_conductance.png","K1"),
                               ("fig_pumping.png","K24"),("fig_FOM_vs_velocity.png","A47")]):
    if os.path.exists(img):
        im=XLImage(img); im.width=im.width*0.7; im.height=im.height*0.7; ws.add_image(im,cell)

wb.save("Sinter_vs_Microchannel_comparison.xlsx")
print("saved Sinter_vs_Microchannel_comparison.xlsx")
